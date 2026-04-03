#!/usr/bin/env python3
"""
Extract receipt data from EPL magazine tracking spreadsheets.
Outputs prisma/seed-receipts.json for import into the app database.

Usage: python3 prisma/extract-receipts.py
"""

import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Canonical magazine name mapping
# ---------------------------------------------------------------------------

# Map spreadsheet names (after stripping issue counts) to canonical seed names.
# Keys are UPPER-CASED for matching.
NAME_MAP = {
    # Main 2025 / 2026
    "AARP BULLETIN": "AARP Bulletin",
    "AARP MAGAZINE": "AARP The Magazine",
    "ALLRECIPES": "All Recipes Magazine",
    "AMERICAN ASSOCIATION OF RETIRED PERSONS MEMBERSHIP": "American Association of Retired Persons Membership",
    "AMERICAN ASSOCIATION OF RETIRED PERSONS": "American Association of Retired Persons Membership",
    "ANANDA VIKATAN - INDIA": "Ananda Vikatan",
    "ANANDA VIKATAN": "Ananda Vikatan",
    "ARCHITECTURAL DIGEST": "Architectural Digest",
    "ARTISTS MAGAZINE": "Artists Magazine",
    "ASK": "Ask",
    "BABY BUG": "Babybug",
    "ASTRONOMY": "Astronomy",
    "ASTRONOMY - PRINT + ONLINE": "Astronomy",
    "ATLANTIC MONTHLY": "Atlantic Monthly",
    "ATLANTIC": "Atlantic Monthly",
    "BABYBUG": "Babybug",
    "BETTER HOMES & GARDEN": "Better Homes and Gardens",
    "BETTER HOMES AND GARDENS": "Better Homes and Gardens",
    "BEANZ": "Beanz",
    "BEANZ - PRINT + ONLINE": "Beanz",
    "BLOOMBERG BUSINESSWEEK": "Bloomberg Businessweek",
    "BLOOMBERG BUSINESS": "Bloomberg Businessweek",
    "BLOOMBERG  BUSINESSWEEK": "Bloomberg Businessweek",
    "BON APPETIT": "Bon Appetit",
    "BOOKLIST": "Booklist",
    "CAR AND DRIVER": "Car and Driver",
    "CHAMPAK(GUAJARATI)": "Champak (Gujarati Edition)",
    "CHAMPAK(GUJARATI)": "Champak (Gujarati Edition)",
    "CHAMPAK(HINDI)": "Champak (Hindi Edition)",
    "CHAMPAK(TAMIL)": "Champak (Tamil Edition)",
    "CHAMPAK(TELUGU)": "Champak (Telugu Edition)",
    "CHINA TODAY - CHINESE ED": "China Today - Chinese Ed",
    "CHIRP": "Chirp",
    "CHITRALEKHA(GUJARATI)": "Chitralekha (Gujarati)",
    "CONSUMER REPORTS": "Consumer Reports",
    "CONSUMER REPORTS BUYING GUIDE": "Consumer Reports Buying Guide",
    "CONSUMER REPORTS BUYING GUIDE - ONLINE": "Consumer Reports Buying Guide",
    "CONSUMER REPORTS ON HEALTH": "Consumer Reports on Health",
    "COOK'S COUNTY": "Cook's Country",
    "COOKS COUNTRY": "Cook's Country",
    "COOKS ILLUSTRATED": "Cooks Illustrated",
    "COSMOPOLITAN": "Cosmopolitan",
    "COUNTRY LIVING": "Country Living",
    "COUNTRY LIVING - NY": "Country Living",
    "CROSSWORD PUZZLE ONLY": "Crossword Puzzles Only",
    "CROSSWORD PUZZLES ONLY": "Crossword Puzzles Only",
    "DISCOVER": "Discover",
    "DISCOVER - PRINT + ONLINE": "Discover",
    "ECONOMIST - US ED": "Economist",
    "ECONOMIST": "Economist",
    "ECONOMIST - PRINT + ONLINE + DIGITAL ED - SINGLE USER": "Economist",
    "ECONOMIST - PRINT + ONLINE +      DIGITAL ED - SINGLE USER": "Economist",
    "ELLE DECOR": "Elle Decor",
    "ELLE D\u00c9COR": "Elle Decor",
    "ENTREPRENEUR": "Entrepreneur",
    "ESQUIRE": "Esquire",
    "ESSENCE": "Essence",
    "ESSENCE - 1 YEAR": "Essence",
    "FAMILY HANDYMAN": "Family Handyman",
    "FAMILY TREE MAGAZINE": "Family Tree Magazine",
    "FINE GARDENING": "Fine Gardening",
    "FIRST": "First for Women",
    "FIRST FOR WOMEN": "First for Women",
    "FOOD NETWORK": "Food Network Magazine",
    "FOOD NETWORK MAGAZINE": "Food Network Magazine",
    "FOOD & WINE": "Food & Wine",
    "FORBES": "Forbes",
    "FORTUNE": "Fortune - Domestic Ed",
    "FORTUNE - DOMESTIC ED": "Fortune - Domestic Ed",
    "FUN FOR KIDZ": "Fun for Kidz",
    "GOLF DIGEST": "Golf Digest",
    "GOOD HOUSEKEEPING": "Good Housekeeping",
    "GRIHSHOBHA(GUJARATI)": "GrihShobha (Gujarati)",
    "GRIHSHOBHA(HINDI)": "GrihShobha (Hindi)(IND)",
    "GRIHSHOBHA(TAMIL)": "GrihShobha (Tamil)",
    "GRIHSHOBHA(TELUGU)": "GrihShobha (Telugu)",
    "GQ - US ED": "GQ - US Edition",
    "GQ - US EDITION": "GQ - US Edition",
    "HARPERS BAZAAR": "Harpers Bazaar",
    "HARVARD BUSINESS REVIEW": "Harvard Business Review",
    "HARVARD BUSINESS REVIEW - P + O + D": "Harvard Business Review",
    "HARVARD HEALTH LETTER": "Harvard Health Letter",
    "HARVARD HEALTH LETTER - PRINT + ONLINE": "Harvard Health Letter",
    "HGTV MAGAZINE": "HGTV Magazine",
    "HIGHLIGHTS": "Highlights for Children",
    "HIGHLIGHTS FOR CHILDREN": "Highlights for Children",
    "HIGHLIGHTS HIGH 5": "Highlights High Five",
    "HIGHLIGHTS HIGH FIVE": "Highlights High Five",
    "HOCKEY NEWS": "Hockey News",
    "HOCKEY NEWS - CANADA": "Hockey News",
    "HOME & DESIGN MAGAZINE": "Home & Design Magazine",
    "HOUSE BEAUTIFUL": "House Beautiful",
    "HUMPTY DUMPTY": "Humpty Dumpty Magazine",
    "HUMPTY DUMPTY MAGAZINE": "Humpty Dumpty Magazine",
    "INC": "Inc",
    "INC -": "Inc",
    "INC 500": "Inc 500",
    "INC 5000 FALL 2025": "Inc 500",
    "KIPLINGERS PERSONAL FINANCE": "Kiplingers Personal Finance",
    "KIRKUS REVIEWS": "Kirkus Reviews",
    "KIRKUS REVIEWS - PRINT + ONLINE - SINGLE USER": "Kirkus Reviews",
    "LADYBUG": "Ladybug",
    "LIBRARY JOURNAL": "Library Journal",
    "MAD": "MAD",
    "MAKE": "Make",
    "MAKE : TECHNOLOGY ON YOUR TIME": "Make",
    "MAGNOLIA": "Magnolia Journal",
    "MENS HEALTH": "Mens Health",
    "MENS HEALTH - PA": "Mens Health",
    "MINDFUL": "Mindful",
    "MINDFUL - ELECTRONIC MAIL": "Mindful",
    "MOTHER EARTH NEWS": "Mother Earth News",
    "MUSE": "Muse",
    "NATIONAL GEOGRAPHIC": "National Geographic",
    "NATIONAL GEOGRAPHIC HISTORY": "National Geographic History",
    "NATIONAL GEO. KIDS": "National Geographic Kids",
    "NATIONAL GEOGRAPHIC - KIDS": "National Geographic Kids",
    "NATIONAL GEOGRAPHIC KIDS": "National Geographic Kids",
    "NAT GEO LITTLE KIDS": "National Geographic Little Kids",
    "NATIONAL GEOGRAPHIC - LITTLE KIDS": "National Geographic Little Kids",
    "NATIONAL GEOGRAPHIC LITTLE KIDS": "National Geographic Little Kids",
    "NATIONAL GEOGRAPHIC SOCIETY MEMBERSHIP": "National Geographic",
    "NEW JERSEY MONTHLY": "New Jersey Monthly",
    "NEW YORK": "New York",
    "NEW YORKER": "New Yorker",
    "OUT": "Out",
    "PASTEL JOURNAL": "Pastel Journal",
    "PEOPLE": "People",
    "PIONEER WOMAN": "Pioneer Woman",
    "POETRY": "Poetry",
    "POETS & WRITERS MAGAZINE": "Poets & Writers Magazine",
    "POPULAR MECHANICS": "Popular Mechanics",
    "POPULAR MECHANICS - ENGLISH ED": "Popular Mechanics",
    "PREVENTION": "Prevention",
    "PREVENTION - PA": "Prevention",
    "PSYCHOLOGY TODAY": "Psychology Today",
    "PUBLISHERS WEEKLY": "Publishers Weekly",
    "PUBLISHERS WEEKLY - PRINT +ONLINE + DIG ED": "Publishers Weekly",
    "QUILTERS WORLD": "Quilters World",
    "RANGER RICK": "Ranger Rick",
    "RANGER RICK - AMERICAN ED": "Ranger Rick",
    "RANGER RICK - AMERICAN EDITION": "Ranger Rick",
    "RANGER RICK JR": "Ranger Rick Jr",
    "RANGER RICK JR.": "Ranger Rick Jr",
    "READERS DIGEST - LARGE PRINT FOR EASIER READING": "Readers Digest - Large Print",
    "READERS DIGEST - LARGE PRINT": "Readers Digest - Large Print",
    "READERS DIGEST - US ED(1)": "Readers Digest - US Ed",
    "READERS DIGEST - US ED": "Readers Digest - US Ed",
    "REAL SIMPLE": "Real Simple",
    "RUNNERS WORLD": "Runners World",
    "SARAS SALIL(HINDI)": "Saras Salil (Hindi Edition)",
    "SARAS SALIL": "Saras Salil (Hindi Edition)",
    "SARITA(HINDI)": "Sarita (Hindi)",
    "SARITA": "Sarita (Hindi)",
    "SCHOOL LIBRARY JOURNAL": "School Library Journal",
    "SCIENCE NEWS": "Science News",
    "SCIENTIFIC AMERICAN": "Scientific American",
    "SCOUT LIFE": "Scout Life",
    "SERIES MADE SIMPLE": "Series Made Simple",
    "SMITHSONIAN": "Smithsonian",
    "SPIDER": "Spider",
    "SPIDER?": "Spider",
    "SPORTS ILLUSTRATED": "Sports Illustrated",
    "SPORTS ILLUSTRATED FOR KIDS": "Sports Illustrated Kids",
    "SPORTS ILLUSTRATED KIDS": "Sports Illustrated Kids",
    "SUPERMAN": "Superman",
    "SUPERMAN?": "Superman",
    "SWATI SAPARIVARA PATRIKA(TELUGU)": "Swati Saparivara Patrika (Telugu)",
    "SWATI SAPARIVARA PATRIKA": "Swati Saparivara Patrika (Telugu)",
    "TASTE OF HOME": "Taste of Home",
    "THREADS": "Threads",
    "TIME MAGAZINE": "Time Magazine",
    "TIME MAGAZINE - DOMESTIC ED": "Time Magazine",
    "TOWN & COUNTRY": "Town & Country",
    "TRAVEL & LEISURE": "Travel & Leisure",
    "US WEEKLY": "US Weekly",
    "VANITY FAIR - AMERICAN ED": "Vanity Fair - American Ed",
    "VEGNEWS MAGAZINE": "VegNews Magazine",
    "VERANDA": "Veranda",
    "VOGUE": "Vogue",
    "THE WEEK - US ED": "The Week - US Edition",
    "THE WEEK - US EDITION": "The Week - US Edition",
    "THE WEEK JUNIOR": "The Week Junior",
    "THE WEEK JUNIOR -": "The Week Junior",
    "WIRED": "Wired",
    "WOMAN'S HEALTH": "Womens Health",
    "WOMEN'S HEALTH": "Womens Health",
    "WOMENS HEALTH": "Womens Health",
    "WOMENS HEALTH - PA": "Womens Health",
    "ZOOKBOOK": "Zoobooks",
    "ZOO BOOK": "Zoobooks",
    "ZOOBOOKS": "Zoobooks",
}

MONTH_NAMES = {
    "JAN": 1, "JANUARY": 1,
    "FEB": 2, "FEBRUARY": 2,
    "MAR": 3, "MARCH": 3,
    "APR": 4, "APRIL": 4,
    "MAY": 5,
    "JUN": 6, "JUNE": 6,
    "JUL": 7, "JULY": 7,
    "AUG": 8, "AUGUST": 8,
    "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10,
    "NOV": 11, "NOVEMBER": 11,
    "DEC": 12, "DECEMBER": 12,
}

SEASON_MAP = {
    "SPRING": (3, 1),
    "SUMMER": (6, 1),
    "FALL": (9, 1),
    "AUTUMN": (9, 1),
    "WINTER": (12, 1),
    "HOLIDAY": (12, 1),
    "LATE FALL": (11, 1),
    "FALL/WINTER": (9, 1),
}


# ---------------------------------------------------------------------------
# Name normalization
# ---------------------------------------------------------------------------

def normalize_main_name(raw: str) -> str:
    """Strip issue count suffix like '(10)' and trailing whitespace from Main spreadsheet names.
    Preserves language tags like (Hindi), (Gujarati), etc."""
    raw = raw.strip()
    # Only strip trailing parentheticals that look like issue counts (digits)
    # or known non-name suffixes like (semi monthly)
    raw = re.sub(r'\s*\(\d+\)\s*$', '', raw)
    raw = re.sub(r'\s*\(semi monthly\)\s*$', '', raw, flags=re.IGNORECASE)
    return raw.strip()


def normalize_ne_name(raw: str) -> str:
    """Strip EBSCO descriptors after /**/ and other suffixes from NE spreadsheet names."""
    raw = raw.strip()
    # Strip everything after /**/
    if '/**/' in raw:
        raw = raw[:raw.index('/**/')].strip()
    elif '/**' in raw:
        raw = raw[:raw.index('/**')].strip()
    # Strip /FOR .../ and /SURFACE MAIL/ etc
    raw = re.sub(r'\s*/[A-Z ]+/\s*', ' ', raw).strip()
    # Strip /FORMERLY/ ... pattern
    raw = re.sub(r'\s*/FORMERLY/.*', '', raw).strip()
    # Strip - PRINT + ONLINE etc suffixes for lookup
    return raw.strip()


def resolve_name(raw: str, source: str) -> str | None:
    """Resolve a raw spreadsheet name to canonical seed name."""
    if source == "NE":
        cleaned = normalize_ne_name(raw)
    else:
        cleaned = normalize_main_name(raw)

    if not cleaned:
        return None

    # First try the full cleaned name (preserves parenthetical language tags)
    key_full = cleaned.upper().strip()
    if key_full in NAME_MAP:
        return NAME_MAP[key_full]

    # Strip trailing parenthetical (issue counts) and try again
    key = re.sub(r'\s*\(\d+\)\s*$', '', key_full).strip()
    if key != key_full and key in NAME_MAP:
        return NAME_MAP[key]

    # Also try stripping ALL trailing parentheticals
    key_bare = re.sub(r'\s*\([^)]*\)\s*$', '', key_full).strip()
    if key_bare != key_full and key_bare in NAME_MAP:
        return NAME_MAP[key_bare]

    # For names with multiple parentheticals, strip one at a time
    key2 = re.sub(r'\s*\([^)]*\)\s*$', '', key_bare).strip()
    if key2 != key_bare and key2 in NAME_MAP:
        return NAME_MAP[key2]

    # Try stripping common suffixes
    for suffix in [" - INDIA", " - CANADA", " - NY", " - PA",
                   " - PRINT + ONLINE", " - PRINT +ONLINE + DIG ED",
                   " - P + O + D", " - ONLINE", " - 1 YEAR",
                   " - AMERICAN EDITION", " - ENGLISH ED"]:
        stripped = key.replace(suffix.upper(), "").strip()
        if stripped in NAME_MAP:
            return NAME_MAP[stripped]

    # Try prefix matching (for long NE names with extra descriptors)
    for map_key, canonical in NAME_MAP.items():
        if key.startswith(map_key) and len(map_key) > 3:
            return canonical

    # Skip known junk/unmapped entries (Cricket is not tracked in this system)
    SKIP_NAMES = {"]", "", "CRICKET"}
    if key in SKIP_NAMES:
        return None

    return None


# ---------------------------------------------------------------------------
# Cell value parsing
# ---------------------------------------------------------------------------

def parse_month_year_from_sheet(sheet_name: str) -> tuple[int, int]:
    """Parse month and year from NE sheet name like 'JAN 25', 'MAR26', ' JUNE 25'."""
    s = sheet_name.strip().upper()
    m = re.match(r'([A-Z]+)\s*(\d{2,4})', s)
    if m:
        month_str, year_str = m.group(1), m.group(2)
        month = MONTH_NAMES.get(month_str, 0)
        year = int(year_str)
        if year < 100:
            year += 2000
        return month, year
    return 0, 0


def parse_cell(value, col_month: int, col_year: int) -> list[str]:
    """
    Parse a cell value into a list of date strings (YYYY-MM-DD).
    col_month: 1-12, the month this column represents.
    col_year: the year context (2025 or 2026).
    """
    if value is None:
        return []

    # Handle datetime objects (from NE sheets via openpyxl)
    if isinstance(value, datetime):
        d = value.date()
        # Sanity: skip obviously wrong dates
        if d.year < 2024 or d.year > 2027:
            return []
        return [d.isoformat()]

    if isinstance(value, date):
        if value.year < 2024 or value.year > 2027:
            return []
        return [value.isoformat()]

    s = str(value).strip()
    if not s or s.lower() == 'nan':
        return []

    # Strip *2 multiplier (NE: means 2 copies, not 2 receipts)
    s = re.sub(r'\*\d+$', '', s).strip()

    # Skip certain non-receipt values
    skip_patterns = [
        r'^UNTIL\b', r'^DISPLAY UNTIL\b', r'^SPECIAL\b', r'^IDEA BOOK',
        r'^PREVIEW\b', r'^THE CUT\b', r'^`$',
    ]
    for pat in skip_patterns:
        if re.match(pat, s, re.IGNORECASE):
            return []

    results = []

    # X or x -> 1st of the column's month
    if s.upper() == 'X':
        return [date(col_year, col_month, 1).isoformat()]

    # "First, Second" (bi-weekly) -> 1st and 15th
    if re.match(r'^First,?\s*Second', s, re.IGNORECASE):
        return [
            date(col_year, col_month, 1).isoformat(),
            date(col_year, col_month, 15).isoformat(),
        ]

    # "First," alone -> 1st of month
    if re.match(r'^First,?\s*$', s, re.IGNORECASE):
        return [date(col_year, col_month, 1).isoformat()]

    # Handle compound cells: split on comma, process each part
    # But first, handle some special compound cases

    # "X,WINTER 2025" or "vol79 no. 06 2026,X"
    # Split and process each part
    parts = _split_cell_parts(s)

    for part in parts:
        part = part.strip()
        if not part:
            continue
        parsed = _parse_single_part(part, col_month, col_year)
        results.extend(parsed)

    return results


def _split_cell_parts(s: str) -> list[str]:
    """Split a cell value into individual parts to parse separately."""
    # Some cells have comma-separated entries. We need to be careful about
    # entries like "DEC 2024/JAN 2025" or "12/26 2025-1/02 2026" that contain
    # slashes and shouldn't be split wrong.

    # Split on commas, but rejoin parts that look like they belong together
    raw_parts = s.split(',')
    parts = []
    i = 0
    while i < len(raw_parts):
        p = raw_parts[i].strip()
        # If this part starts a cross-year entry, it might continue
        parts.append(p)
        i += 1
    return parts


def _parse_single_part(s: str, col_month: int, col_year: int) -> list[str]:
    """Parse a single part of a cell value."""
    s = s.strip()
    if not s:
        return []

    # Strip *N suffix
    s = re.sub(r'\*\d+$', '', s).strip()

    # X or x
    if s.upper() == 'X':
        return [date(col_year, col_month, 1).isoformat()]

    # Skip noise
    skip_patterns = [
        r'^UNTIL\b', r'^DISPLAY UNTIL\b', r'^SPECIAL\b', r'^IDEA BOOK',
        r'^PREVIEW\b', r'^THE CUT\b', r'^`$', r'^\(2 COPIES\)$',
        r'^2 COPIES$',
    ]
    for pat in skip_patterns:
        if re.match(pat, s, re.IGNORECASE):
            return []

    # Full date like "2025-01-25 00:00:00" (string version of datetime)
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2024 <= y <= 2027:
            return [date(y, mo, d).isoformat()]
        return []

    # Full date like "1/13/2025"
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2024 <= y <= 2027:
            return [date(y, mo, d).isoformat()]
        return []

    # Season with year: "SPRING 2025", "SUMMER 2025", "FALL 2025", "WINTER 2025"
    # Also: "WINTER 2024/2025", "WINTER 2025/2026", "WINTER 2025/26"
    m = re.match(r'^(SPRING|SUMMER|FALL|AUTUMN|WINTER|HOLIDAY|LATE FALL|FALL/WINTER)\s*(\d{4})?(?:/\d{2,4})?$', s, re.IGNORECASE)
    if m:
        season = m.group(1).upper()
        year_str = m.group(2)
        if season in SEASON_MAP:
            sm, sd = SEASON_MAP[season]
            sy = int(year_str) if year_str else col_year
            return [date(sy, sm, sd).isoformat()]

    # Season without year: "SPRING", "SUMMER", "WINTER", "HOLIDAY", "FALL 2025"
    # Also "SPR 25", "SPRING 25", "SUMMER 25"
    m = re.match(r'^(SPRING|SPR|SUMMER|FALL|AUTUMN|WINTER|HOLIDAY|LATE FALL|FALL/WINTER)\s*(\d{2})?$', s, re.IGNORECASE)
    if m:
        season_raw = m.group(1).upper()
        year_suffix = m.group(2)
        season_key = {"SPR": "SPRING"}.get(season_raw, season_raw)
        if season_key in SEASON_MAP:
            sm, sd = SEASON_MAP[season_key]
            sy = (2000 + int(year_suffix)) if year_suffix else col_year
            return [date(sy, sm, sd).isoformat()]

    # "HOLIDAY ISSUE 2025" or "HOLIDAY 2025"
    m = re.match(r'^HOLIDAY\s+(?:ISSUE\s+)?(\d{4})$', s, re.IGNORECASE)
    if m:
        return [date(int(m.group(1)), 12, 1).isoformat()]

    # "INC 5000 FALL 2025" -> special
    m = re.match(r'^INC 5000\s+(FALL|SPRING|SUMMER|WINTER)\s+(\d{4})$', s, re.IGNORECASE)
    if m:
        season = m.group(1).upper()
        year = int(m.group(2))
        if season in SEASON_MAP:
            sm, sd = SEASON_MAP[season]
            return [date(year, sm, sd).isoformat()]

    # "SPECIAL EDITION 2025" -> col_month 1st
    m = re.match(r'^SPECIAL\s+EDITION\s+(\d{4})$', s, re.IGNORECASE)
    if m:
        return [date(int(m.group(1)), col_month, 1).isoformat()]

    # Cross-year: "DEC 2024/JAN 2025", "DEC 2025/ JAN 2026", "DEC/JAN 2025"
    # "DEC 2024 / JAN 2025", "DEC 2024- JAN 2025", "DEC 2025/JAN2026"
    m = re.match(
        r'^(?:DEC(?:EMBER)?)\s*(\d{4})?\s*[-/&]\s*(?:JAN(?:UARY)?)\s*(\d{2,4})?',
        s, re.IGNORECASE
    )
    if m:
        year_str = m.group(2) or m.group(1)
        if year_str:
            y = int(year_str)
            if y < 100:
                y += 2000
            return [date(y, 1, 1).isoformat()]
        else:
            # DEC/JAN with no year - use col_year, Jan of next year if col is Dec
            if col_month == 12:
                return [date(col_year + 1, 1, 1).isoformat()]
            elif col_month == 1:
                return [date(col_year, 1, 1).isoformat()]
            return [date(col_year, 1, 1).isoformat()]

    # "DEC/JAN" standalone (NE sheets)
    m = re.match(r'^DEC[-/]JAN\s*(\d{2,4})?$', s, re.IGNORECASE)
    if m:
        year_str = m.group(1)
        if year_str:
            y = int(year_str)
            if y < 100:
                y += 2000
            return [date(y, 1, 1).isoformat()]
        # Use context: if in Jan sheet, Jan of that year
        if col_month <= 1:
            return [date(col_year, 1, 1).isoformat()]
        return [date(col_year + 1, 1, 1).isoformat()]

    # "Nov-Dec", "Dec-Jan" (NE DEC 25 sheet)
    m = re.match(r'^([A-Z]+)\s*[-/]\s*([A-Z]+)$', s, re.IGNORECASE)
    if m:
        m1 = MONTH_NAMES.get(m.group(1).upper().strip())
        m2 = MONTH_NAMES.get(m.group(2).upper().strip())
        if m1 and m2:
            return [date(col_year, m1, 1).isoformat()]

    # Month pair: "JAN/FEB", "MAR/APR", "MAY/.JUN", "JUN/JUL", "AUG/SEP", etc
    m = re.match(r'^([A-Z]+)\s*/?\.?\s*([A-Z]+)\s*(\d{2,4})?\s*$', s, re.IGNORECASE)
    if m:
        m1_str = m.group(1).upper().strip()
        m2_str = m.group(2).upper().strip()
        year_str = m.group(3)
        m1 = MONTH_NAMES.get(m1_str)
        m2 = MONTH_NAMES.get(m2_str)
        if m1 and m2:
            y = col_year
            if year_str:
                y = int(year_str)
                if y < 100:
                    y += 2000
            return [date(y, m1, 1).isoformat()]

    # "DEC 21ST - JAN 3RD 2025,..." or "DEC 27 2024- JAN 3RD 2025,..."
    # "DEC 30 - JAN 12 2025,..."
    m = re.match(r'^DEC\s+\d+(?:ST|ND|RD|TH)?\s*(?:\d{4})?\s*[-&]\s*JAN\s+\d+(?:ST|ND|RD|TH)?\s*(\d{4})?', s, re.IGNORECASE)
    if m:
        y = int(m.group(1)) if m.group(1) else col_year
        return [date(y, 1, 1).isoformat()]

    # "DEC 30 2024 & JAN 6 2025,..." -> two entries
    m = re.match(r'^DEC\s+(\d+)\s+(\d{4})\s*[&]\s*JAN\s+(\d+)\s*(\d{4})?', s, re.IGNORECASE)
    if m:
        y2 = int(m.group(4)) if m.group(4) else col_year
        return [date(y2, 1, int(m.group(3))).isoformat()]

    # "12/26 2025-1/02 2026" or "12/20 205-1/2 2026" (typo for 2025)
    m = re.match(r'^12/(\d+)\s+\d{3,4}\s*-\s*1/(\d+)\s+(\d{4})', s)
    if m:
        y = int(m.group(3))
        return [date(y, 1, int(m.group(2))).isoformat()]

    # "12/29 2025 -1/5 2026,..." cross-year prefix
    m = re.match(r'^12/\d+\s+\d{4}\s*-\s*1/(\d+)\s+(\d{4})', s)
    if m:
        y = int(m.group(2))
        return [date(y, 1, int(m.group(1))).isoformat()]

    # Date range: "1/11-17" or "1/11-1/17" or "6/7-6/13" or "3/15-21"
    # "12/30-1/12" (cross-month)
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*-\s*(?:(\d{1,2})/)?(\d{1,2})$', s)
    if m:
        m1 = int(m.group(1))
        d1 = int(m.group(2))
        # Use first date of the range
        y = col_year
        # Handle cross-year: month 12 range crossing to month 1
        if m1 == 12 and col_month == 1:
            y = col_year - 1
        try:
            return [date(y, m1, d1).isoformat()]
        except ValueError:
            return []

    # "13-19 Sep" or "12-19 May" or "8-14Nov" or "15-21Nov" (day range with month name)
    m = re.match(r'^(\d{1,2})\s*-\s*\d{1,2}\s*([A-Za-z]+)$', s)
    if m:
        d1 = int(m.group(1))
        month_str = m.group(2).upper().strip()
        mo = MONTH_NAMES.get(month_str)
        if mo:
            try:
                return [date(col_year, mo, d1).isoformat()]
            except ValueError:
                return []

    # "June 7-13" (month name then day range)
    m = re.match(r'^([A-Za-z]+)\s+(\d{1,2})\s*-\s*(\d{1,2})$', s)
    if m:
        month_str = m.group(1).upper().strip()
        d1 = int(m.group(2))
        mo = MONTH_NAMES.get(month_str)
        if mo:
            try:
                return [date(col_year, mo, d1).isoformat()]
            except ValueError:
                return []

    # Date with copy note: "5/12(2 COPIES)"
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*\(.*\)$', s)
    if m:
        mo = int(m.group(1))
        d = int(m.group(2))
        try:
            return [date(col_year, mo, d).isoformat()]
        except ValueError:
            return []

    # Cross-month range with year: "12/29-1/5 2026" or "12/01-12/14,12/15-12/28,12/29-1/5 2026"
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*-\s*(\d{1,2})/(\d{1,2})\s+(\d{4})$', s)
    if m:
        m1, d1 = int(m.group(1)), int(m.group(2))
        m2, d2 = int(m.group(3)), int(m.group(4))
        y = int(m.group(5))
        # Use first date, but year context depends on the month
        if m1 == 12:
            return [date(y - 1, m1, d1).isoformat()]
        return [date(y, m1, d1).isoformat()]

    # Date range with hyphen instead of slash: "1/31-2-06" (typo: should be 1/31-2/06)
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*-\s*(\d{1,2})-(\d{1,2})$', s)
    if m:
        m1 = int(m.group(1))
        d1 = int(m.group(2))
        try:
            return [date(col_year, m1, d1).isoformat()]
        except ValueError:
            return []

    # Simple date: "M/D" like "1/13" or "10/01"
    m = re.match(r'^(\d{1,2})/(\d{1,2})$', s)
    if m:
        mo = int(m.group(1))
        d = int(m.group(2))
        try:
            return [date(col_year, mo, d).isoformat()]
        except ValueError:
            return []

    # "2/17 & 2/24" or "12/30 & 1/6"
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*&\s*(\d{1,2})/(\d{1,2})$', s)
    if m:
        results = []
        m1, d1 = int(m.group(1)), int(m.group(2))
        m2, d2 = int(m.group(3)), int(m.group(4))
        try:
            y1 = col_year
            if m1 == 12 and col_month <= 2:
                y1 = col_year - 1
            results.append(date(y1, m1, d1).isoformat())
        except ValueError:
            pass
        try:
            results.append(date(col_year, m2, d2).isoformat())
        except ValueError:
            pass
        return results

    # "VOL.78 NO.05" or "VOL 79 NO. 04" or "vol79 no.03"
    m = re.match(r'^VOL[.: ]?\s*\d+\s*NO[.: ]\s*[O0]?\d+', s, re.IGNORECASE)
    if m:
        return [date(col_year, col_month, 1).isoformat()]

    # "vol.79 no. 06 2026,X" compound
    m = re.match(r'^VOL[.: ]?\s*\d+\s*NO[.: ]\s*\d+\s*\d{4}', s, re.IGNORECASE)
    if m:
        return [date(col_year, col_month, 1).isoformat()]

    # Single month name: "JAN", "FEB", "MAR" etc
    m = re.match(r'^([A-Z]+)$', s, re.IGNORECASE)
    if m:
        mo = MONTH_NAMES.get(m.group(1).upper())
        if mo:
            return [date(col_year, mo, 1).isoformat()]

    # "APR/MAY 25" -> month pair with year
    m = re.match(r'^([A-Z]+)\s*/\s*([A-Z]+)\s+(\d{2,4})$', s, re.IGNORECASE)
    if m:
        m1 = MONTH_NAMES.get(m.group(1).upper())
        m2 = MONTH_NAMES.get(m.group(2).upper())
        y = int(m.group(3))
        if y < 100:
            y += 2000
        if m1:
            return [date(y, m1, 1).isoformat()]

    # "MAY/JUNE 25" or "may/june"
    m = re.match(r'^([A-Z]+)\s*/\s*([A-Z]+)$', s, re.IGNORECASE)
    if m:
        m1 = MONTH_NAMES.get(m.group(1).upper())
        m2 = MONTH_NAMES.get(m.group(2).upper())
        if m1 and m2:
            return [date(col_year, m1, 1).isoformat()]

    # "SPRINF 2026" (typo for SPRING)
    m = re.match(r'^SPRINF\s+(\d{4})$', s, re.IGNORECASE)
    if m:
        return [date(int(m.group(1)), 3, 1).isoformat()]

    # "WINTER 2025/2026" or "WINTER 2025/26"
    m = re.match(r'^WINTER\s+(\d{4})/(\d{2,4})$', s, re.IGNORECASE)
    if m:
        y = int(m.group(1))
        return [date(y, 12, 1).isoformat()]

    # "6/2----6/5" (NE typo)
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*-+\s*(\d{1,2})/(\d{1,2})$', s)
    if m:
        mo = int(m.group(1))
        d = int(m.group(2))
        try:
            return [date(col_year, mo, d).isoformat()]
        except ValueError:
            return []

    # "4/4,4/11,4/18,4/25" inside a part that somehow wasn't split
    # (shouldn't happen since we split on commas, but just in case)
    m = re.match(r'^(\d{1,2})/(\d{1,2})$', s)
    if m:
        mo, d = int(m.group(1)), int(m.group(2))
        try:
            return [date(col_year, mo, d).isoformat()]
        except ValueError:
            return []

    # "7/12 - 7/18" with spaces
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s+-\s+(\d{1,2})/(\d{1,2})$', s)
    if m:
        mo = int(m.group(1))
        d = int(m.group(2))
        try:
            return [date(col_year, mo, d).isoformat()]
        except ValueError:
            return []

    # "11./10" (NE typo for 11/10)
    m = re.match(r'^(\d{1,2})\./(\d{1,2})$', s)
    if m:
        mo = int(m.group(1))
        d = int(m.group(2))
        try:
            return [date(col_year, mo, d).isoformat()]
        except ValueError:
            return []

    # If we get here, we didn't parse it. Log and skip.
    print(f"  WARNING: Could not parse cell value: {repr(s)} (month={col_month}, year={col_year})", file=sys.stderr)
    return []


# ---------------------------------------------------------------------------
# Spreadsheet readers
# ---------------------------------------------------------------------------

def read_main_sheet(filepath: str, year: int) -> list[dict]:
    """Read a Main Library spreadsheet (2025 or 2026 format)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    receipts = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        raw_name = str(row[0].value).strip() if row[0].value else ''
        if not raw_name or raw_name == 'None':
            continue

        canonical = resolve_name(raw_name, "MAIN")
        if canonical is None:
            print(f"  WARNING: No mapping for Main name: {repr(raw_name)}", file=sys.stderr)
            continue

        for col_idx in range(1, 13):  # columns B-M = Jan-Dec
            col_month = col_idx  # 1=Jan, 2=Feb, etc.
            if col_idx >= len(row):
                continue
            cell_val = row[col_idx].value
            if cell_val is None:
                continue

            dates = parse_cell(cell_val, col_month, year)
            for d in dates:
                receipts.append({
                    "magazine": canonical,
                    "branch": "MAIN",
                    "date": d,
                    "notes": str(cell_val).strip()[:80],
                })

    return receipts


def read_ne_sheets(filepath: str) -> list[dict]:
    """Read North Edison spreadsheet (one sheet per month, JAN 25 onwards)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    receipts = []

    target_sheets = [
        'JAN 25', 'FEB 25', 'MAR 25', 'APR 25', 'MAY 25', ' JUNE 25',
        'JULY 25', 'AUG 25', 'SEPT 25', 'OCT 25', 'NOV 25', 'DEC 25',
        'Jan 26', 'FEB 26', 'MAR26',
    ]

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet {repr(sheet_name)} not found in NE workbook", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        sheet_month, sheet_year = parse_month_year_from_sheet(sheet_name)
        if sheet_month == 0:
            print(f"  WARNING: Could not parse month/year from sheet name: {repr(sheet_name)}", file=sys.stderr)
            continue

        # Track seen magazine names to detect the second-copy section
        seen_names = set()

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            raw_name = str(row[0].value).strip() if row[0].value else ''
            if not raw_name or raw_name == 'None':
                continue

            canonical = resolve_name(raw_name, "NE")
            if canonical is None:
                print(f"  WARNING: No mapping for NE name: {repr(raw_name)} (sheet={sheet_name})", file=sys.stderr)
                continue

            # Deduplicate: skip second-copy entries
            if canonical in seen_names:
                continue
            seen_names.add(canonical)

            for col_idx in range(1, min(6, ws.max_column + 1)):
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx].value
                if cell_val is None:
                    continue

                dates = parse_cell(cell_val, sheet_month, sheet_year)
                for d in dates:
                    # Fix dates that are exactly 1 year off (common spreadsheet error)
                    d_year = int(d[:4])
                    d_month = int(d[5:7])
                    if abs(d_year - sheet_year) == 1 and d_month == sheet_month:
                        d = f"{sheet_year}{d[4:]}"
                    receipts.append({
                        "magazine": canonical,
                        "branch": "NORTH",
                        "date": d,
                        "notes": str(cell_val).strip()[:80],
                    })

    return receipts


def read_cb_adult_sheets(filepath: str) -> list[dict]:
    """Read Clara Barton adult magazine spreadsheet (sheets '2025' and '2026').

    Format is the same as Main: TITLE in column A, month columns B-M (Jan-Dec).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    receipts = []

    sheet_years = {"2025": 2025, "2026": 2026}
    for sheet_name, year in sheet_years.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet {repr(sheet_name)} not found in CB adult workbook", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            raw_name = str(row[0].value).strip() if row[0].value else ''
            if not raw_name or raw_name == 'None':
                continue

            canonical = resolve_name(raw_name, "MAIN")
            if canonical is None:
                print(f"  WARNING: No mapping for CB adult name: {repr(raw_name)}", file=sys.stderr)
                continue

            for col_idx in range(1, 13):  # columns B-M = Jan-Dec
                col_month = col_idx  # 1=Jan, 2=Feb, etc.
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx].value
                if cell_val is None:
                    continue

                dates = parse_cell(cell_val, col_month, year)
                for d in dates:
                    receipts.append({
                        "magazine": canonical,
                        "branch": "CB",
                        "date": d,
                        "notes": str(cell_val).strip()[:80],
                    })

    return receipts


def read_cb_childrens_sheet(filepath: str) -> list[dict]:
    """Read Clara Barton children's magazine spreadsheet (sheet 'Sheet1').

    Format is the same as Main: TITLE in column A, month columns B-M (Jan-Dec).
    The sheet covers both 2025 and 2026 data split by a year header row.
    We detect the active year by looking for a row whose first cell contains just
    the year (e.g. "2025", "2026") and update context accordingly.
    If no such row is found, we default to 2025 for columns 1-12 and 2026 for
    any second pass of month columns (columns 14-25 if present).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    receipts = []

    sheet_name = "Sheet1"
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet {repr(sheet_name)} not found in CB children's workbook", file=sys.stderr)
        return receipts

    ws = wb[sheet_name]

    # Detect layout: check header row for year groups
    # The sheet may have two sets of month columns: Jan-Dec 2025 (cols B-M) then Jan-Dec 2026 (cols N-Y)
    # OR it may have a single year column set with a year indicator in the name row.
    # We handle both: first 12 month columns = 2025, next 12 = 2026.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        raw_name = str(row[0].value).strip() if row[0].value else ''
        if not raw_name or raw_name == 'None':
            continue

        # Skip header/year indicator rows (cell is just "2025", "2026", or "TITLE" etc.)
        if re.match(r'^\d{4}$', raw_name) or raw_name.upper() in ('TITLE', 'MAGAZINE', 'NAME'):
            continue

        canonical = resolve_name(raw_name, "MAIN")
        if canonical is None:
            print(f"  WARNING: No mapping for CB children's name: {repr(raw_name)}", file=sys.stderr)
            continue

        # Process first 12 month columns (2025): col indices 1-12
        for col_idx in range(1, 13):
            col_month = col_idx
            if col_idx >= len(row):
                continue
            cell_val = row[col_idx].value
            if cell_val is None:
                continue

            dates = parse_cell(cell_val, col_month, 2025)
            for d in dates:
                receipts.append({
                    "magazine": canonical,
                    "branch": "CB",
                    "date": d,
                    "notes": str(cell_val).strip()[:80],
                })

        # Process second 12 month columns (2026): col indices 13-24 (if present)
        if ws.max_column >= 14:
            for col_idx in range(13, 25):
                col_month = col_idx - 12  # 1=Jan, 2=Feb, etc.
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx].value
                if cell_val is None:
                    continue

                dates = parse_cell(cell_val, col_month, 2026)
                for d in dates:
                    receipts.append({
                        "magazine": canonical,
                        "branch": "CB",
                        "date": d,
                        "notes": str(cell_val).strip()[:80],
                    })

    return receipts


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    base = Path(__file__).parent.parent / "docs"

    print("Reading Main 2025...")
    main_2025 = read_main_sheet(str(base / "Ebsco Magazines for Main 2025.xlsx"), 2025)
    print(f"  -> {len(main_2025)} receipts")

    print("Reading Main 2026...")
    main_2026 = read_main_sheet(str(base / "Ebsco Magazines for Main 2026.xlsx"), 2026)
    print(f"  -> {len(main_2026)} receipts")

    print("Reading NE 2025-2026...")
    ne = read_ne_sheets(str(base / "Ebsco NE 2025-2026 Magazine List.xlsx"))
    print(f"  -> {len(ne)} receipts")

    print("Reading CB adult 2025-2026...")
    cb_adult = read_cb_adult_sheets(str(base / "Ebsco CB 2025-2026 Magazine List.xlsx"))
    print(f"  -> {len(cb_adult)} receipts")

    print("Reading CB children's 2025-2026...")
    cb_childrens = read_cb_childrens_sheet(str(base / "Ebsco CB 2025-2026 Childrens Magazine List.xlsx"))
    print(f"  -> {len(cb_childrens)} receipts")

    all_receipts = main_2025 + main_2026 + ne + cb_adult + cb_childrens

    # Deduplicate: same magazine + branch + date
    seen = set()
    deduped = []
    for r in all_receipts:
        key = (r["magazine"], r["branch"], r["date"])
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    print(f"\nTotal: {len(all_receipts)} raw, {len(deduped)} after dedup")

    # Fix obvious year typos (e.g. "2005" should be "2025")
    for r in deduped:
        y = int(r["date"][:4])
        if y < 2024:
            # Likely a typo: fix year to 2025
            r["date"] = "2025" + r["date"][4:]

    # Filter out receipts before the subscription start date (2025-01-01)
    from datetime import date as date_type
    SUBSCRIPTION_START = date_type(2025, 1, 1)
    deduped = [r for r in deduped if r['date'] >= SUBSCRIPTION_START.isoformat()]
    print(f"After filtering >= {SUBSCRIPTION_START}: {len(deduped)} receipts")

    # Sort by date, then magazine, then branch
    deduped.sort(key=lambda r: (r["date"], r["magazine"], r["branch"]))

    # Write output
    out_path = Path(__file__).parent / "seed-receipts.json"
    with open(out_path, 'w') as f:
        json.dump(deduped, f, indent=2)
    print(f"Wrote {len(deduped)} receipts to {out_path}")

    # Cross-check Main 2025 counts
    print("\n--- Cross-check (Main 2025) ---")
    main_2025_by_mag = {}
    for r in main_2025:
        mag = r["magazine"]
        main_2025_by_mag[mag] = main_2025_by_mag.get(mag, 0) + 1

    checks = {
        "Economist": 37,
        "Atlantic Monthly": 9,
        "People": 43,
        "Bloomberg Businessweek": 12,
        "Bon Appetit": 10,
    }
    for mag, expected in checks.items():
        actual = main_2025_by_mag.get(mag, 0)
        status = "OK" if actual == expected else "MISMATCH"
        print(f"  {mag}: expected={expected}, actual={actual} [{status}]")

    # Print per-branch summary
    print("\n--- Per-branch summary ---")
    branch_counts = {}
    for r in deduped:
        branch_counts[r["branch"]] = branch_counts.get(r["branch"], 0) + 1
    for branch, count in sorted(branch_counts.items()):
        print(f"  {branch}: {count} receipts")

    # Print unique magazines
    magazines = sorted(set(r["magazine"] for r in deduped))
    print(f"\n--- Unique magazines: {len(magazines)} ---")
    for m in magazines:
        count = sum(1 for r in deduped if r["magazine"] == m)
        print(f"  {m}: {count}")


if __name__ == "__main__":
    main()
